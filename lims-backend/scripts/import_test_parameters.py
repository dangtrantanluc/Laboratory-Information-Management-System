"""Import master data CHỈ TIÊU THỬ NGHIỆM từ "BẢNG GIÁ PHÂN TÍCH - 2024.xlsx" (m27).

Mỗi sheet = 1 nhóm nền mẫu. Cột KHÔNG giống nhau giữa các sheet nên map theo TIÊU ĐỀ
(dò dòng header có "CHỈ TIÊU THỬ NGHIỆM"), không hardcode vị trí cột.

Chạy:
  docker exec lims-api python scripts/import_test_parameters.py /path/file.xlsx [--dry-run]

Idempotent: bỏ qua dòng đã tồn tại (cùng matrix + tên + phương pháp, không phân biệt hoa/thường).
"""
import argparse
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation

import openpyxl

sys.path.insert(0, "/app")

from app.db.database import SessionLocal  # noqa: E402
from app.models.department import Department  # noqa: E402
from app.models.sample_flow import TestParameter  # noqa: E402

# Tên sheet → mã nhóm nền mẫu
SHEET_MATRIX = {
    "đất": "soil",
    "nước": "water",
    "phân bón, chế phẩm sinh học": "fertilizer",
    "thức ăn chăn nuôi": "feed",
    "nông sản, thực phẩm": "food",
    "kiểm dịch thực vật": "quarantine",
    "shpt": "molecular",
}

# Sheet phân tử → gán sẵn phòng lab SHPT (phương pháp PCR/định danh SHPT rõ ràng).
# Các nhóm còn lại để TRỐNG cho Phòng nhận mẫu/Quản trị tự gán trên UI.
MATRIX_DEFAULT_DEPT_CODE = {
    "molecular": "LAB-SHPT",
    "quarantine": "LAB-SHPT",
}


def norm(s) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s))
    return re.sub(r"\s+", " ", s).strip()


def norm_key(s) -> str:
    return norm(s).lower()


def parse_price(v):
    if v is None or norm(v) == "":
        return None
    if isinstance(v, (int, float, Decimal)):
        return Decimal(str(v))
    txt = re.sub(r"[^\d.,]", "", str(v)).replace(".", "").replace(",", ".")
    if not txt:
        return None
    try:
        return Decimal(txt)
    except InvalidOperation:
        return None


def parse_int(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    m = re.search(r"\d+", str(v))
    return int(m.group()) if m else None


# Khối ký tên/chân bảng ở cuối sheet — gặp là DỪNG đọc sheet đó (tránh nhập rác
# như "Người lập", tên người ký). Không dùng rule "thiếu giá+phương pháp" vì nhiều
# chỉ tiêu thật (Cordycepin, Mycorrhiza spp., các virus…) cũng trống 2 cột đó.
STOP_MARKERS = (
    "người lập", "người duyệt", "người kiểm tra", "người soát xét",
    "giám đốc", "phó giám đốc", "trưởng phòng", "phụ trách phòng",
    "ký tên", "xác nhận", "ngày ... tháng", "tp. hồ chí minh, ngày",
)


def is_stop_row(text: str) -> bool:
    t = norm_key(text)
    return any(m in t for m in STOP_MARKERS)


# Từ khóa nhận cột theo tiêu đề
HEADER_MAP = [
    ("name", ("chỉ tiêu thử nghiệm", "chỉ tiêu")),
    ("method", ("phương pháp thử nghiệm", "phương pháp")),
    ("unit_price", ("đơn giá",)),
    ("sample_matrix", ("nền mẫu",)),
    ("turnaround_days", ("thời gian",)),
    ("note", ("ghi chú",)),
]


def detect_header(ws):
    """Tìm dòng header + map {field: col_index}. Trả (header_row, mapping)."""
    for r in range(1, min(ws.max_row, 12) + 1):
        cells = {c: norm_key(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)}
        if not any("chỉ tiêu" in v for v in cells.values()):
            continue
        mapping = {}
        for field, keys in HEADER_MAP:
            for c, text in cells.items():
                if not text:
                    continue
                if any(k in text for k in keys):
                    mapping.setdefault(field, c)
                    break
        if "name" in mapping:
            return r, mapping
    return None, {}


def guess_in_charge(ws, row, used_cols):
    """Cột người phụ trách không có tiêu đề — dò ô chứa 'Cô/Thầy/Anh/Chị' ở các cột còn lại."""
    for c in range(1, ws.max_column + 1):
        if c in used_cols:
            continue
        v = norm(ws.cell(row, c).value)
        if v and re.match(r"^(Cô|Thầy|Anh|Chị)\s", v):
            return v
    return None


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path, data_only=True)
    db = SessionLocal()

    # map mã phòng → id
    dept_by_code = {
        d.code: d.id for d in db.execute(__import__("sqlalchemy").select(Department)).scalars().all()
    }

    # đã có trong DB → chống trùng
    existing = set()
    for p in db.query(TestParameter).all():
        existing.add((p.matrix, norm_key(p.name), norm_key(p.method or "")))

    total_new = total_skip = 0
    for ws in wb.worksheets:
        matrix = SHEET_MATRIX.get(norm_key(ws.title))
        if not matrix:
            print(f"  ! bỏ qua sheet lạ: {ws.title!r}")
            continue
        header_row, cmap = detect_header(ws)
        if not header_row:
            print(f"  ! không tìm được header: {ws.title!r}")
            continue

        dept_id = dept_by_code.get(MATRIX_DEFAULT_DEPT_CODE.get(matrix, ""))
        used_cols = set(cmap.values()) | {1}
        new = skip = 0
        last_sample_matrix = None
        order = 0

        for r in range(header_row + 1, ws.max_row + 1):
            name = norm(ws.cell(r, cmap["name"]).value)
            if is_stop_row(name):
                break  # tới khối ký tên → hết dữ liệu của sheet
            if not name or norm_key(name) in ("chỉ tiêu thử nghiệm", "chỉ tiêu"):
                continue
            # bỏ dòng tiêu đề nhóm (không có STT và không có giá) — vẫn giữ nếu có giá
            price = parse_price(ws.cell(r, cmap["unit_price"]).value) if "unit_price" in cmap else None
            method = norm(ws.cell(r, cmap["method"]).value) if "method" in cmap else ""

            # NỀN MẪU của sheet SHPT bị merge → fill down
            sm = norm(ws.cell(r, cmap["sample_matrix"]).value) if "sample_matrix" in cmap else ""
            if sm:
                last_sample_matrix = sm
            sample_matrix = last_sample_matrix if "sample_matrix" in cmap else None

            key = (matrix, norm_key(name), norm_key(method))
            if key in existing:
                skip += 1
                continue
            existing.add(key)
            order += 1

            # dấu "*" cuối tên = chỉ tiêu được công nhận VILAS
            is_accredited = name.rstrip().endswith("*")
            clean_name = name.rstrip().rstrip("*").strip()

            tp = TestParameter(
                matrix=matrix,
                sample_matrix=sample_matrix or None,
                name=clean_name or name,
                method=method or None,
                unit_price=price,
                currency="VND",
                turnaround_days=(
                    parse_int(ws.cell(r, cmap["turnaround_days"]).value)
                    if "turnaround_days" in cmap else None
                ),
                in_charge=guess_in_charge(ws, r, used_cols),
                note=(norm(ws.cell(r, cmap["note"]).value) or None) if "note" in cmap else None,
                department_id=dept_id,
                is_accredited=is_accredited,
                is_active=True,
                sort_order=order,
            )
            db.add(tp)
            new += 1

        total_new += new
        total_skip += skip
        print(f"  {ws.title:32s} matrix={matrix:11s} +{new:4d} mới, {skip:4d} bỏ qua (trùng)")

    if args.dry_run:
        db.rollback()
        print(f"DRY-RUN: sẽ thêm {total_new}, bỏ qua {total_skip}")
    else:
        db.commit()
        print(f"XONG: đã thêm {total_new}, bỏ qua {total_skip}")
    db.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
