"""Import "TỔNG HỢP CÁC HOẠT ĐỘNG NĂM 2024-2025.xlsx" → DB (menu M4).

Đọc theo TỪNG vùng bảng con (header 2 tầng + merge như phân tích trong
docs/PHAN_TICH_SCHEMA_HOAT_DONG_2024_2025.md), LÀM SẠCH dữ liệu bẩn, map tên↔user
(fallback external_name), rồi ghi vào các bảng đã mở rộng ở migration m23.

Chạy:
    # In báo cáo, KHÔNG ghi DB:
    python scripts/import_activities_2024_2025.py --file "docs/TỔNG HỢP...xlsx" --dry-run
    # Ghi thật:
    DATABASE_URL=... python scripts/import_activities_2024_2025.py --file "..." --academic-year 2024-2025

Các helper làm sạch (parse_budget, parse_indexing, ...) là HÀM THUẦN, có test riêng
(app/tests/services/test_activity_import.py).
"""
import argparse
import re
import sys
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import Optional

# ---------------------------------------------------------------- helpers thuần

_MULT = {"nghìn": 1_000, "ngàn": 1_000, "triệu": 1_000_000, "tỷ": 1_000_000_000, "tỉ": 1_000_000_000}


def parse_budget(value) -> Optional[Decimal]:
    """"100 triệu" → 100000000 ; 110000000 → 110000000 ; "" → None."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    s = str(value).strip().lower().replace(",", "").replace(".", "")
    m = re.search(r"([\d]+(?:[.,]?\d+)?)\s*(nghìn|ngàn|triệu|tỷ|tỉ)?", s)
    if not m:
        return None
    try:
        num = Decimal(m.group(1))
    except InvalidOperation:
        return None
    unit = m.group(2)
    return num * _MULT.get(unit, 1) if unit else num


def parse_year_range(value) -> tuple[Optional[int], Optional[int]]:
    """"2023-2025" → (2023, 2025) ; "2025" → (2025, 2025) ; "" → (None, None)."""
    if not value:
        return None, None
    yrs = re.findall(r"(19|20)\d{2}", str(value))
    yrs = [int(str(value)[m.start():m.start()+4]) for m in re.finditer(r"(?:19|20)\d{2}", str(value))]
    if not yrs:
        return None, None
    return yrs[0], yrs[-1]


def parse_int(value) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def parse_indexing(scie_cell, scopus_cell, aci_cell) -> dict:
    """Chuẩn hoá 3 ô chỉ mục bẩn ("SCIE","SCIE/SSCI","SCOPUS","X","x",…) → 4 cờ boolean."""
    def _t(x):
        return str(x).strip().upper() if x not in (None, "") else ""
    a, b, c = _t(scie_cell), _t(scopus_cell), _t(aci_cell)
    blob = f"{a} {b} {c}"
    return {
        "is_scie": "SCIE" in a,
        "is_ssci": "SSCI" in a,
        # Scopus có thể nằm ở cột SCIE (bẩn) hoặc cột Scopus (đánh dấu X)
        "is_scopus": "SCOPUS" in blob or b in ("X", "X.", "✓"),
        "is_aci": "ACI" in blob or c in ("X", "X.", "✓"),
    }


def parse_author_role(value) -> tuple[Optional[str], bool]:
    """"Tác giả liên hệ" → (corresponding, True) ; "ĐTG" → (co, False) ; "TG" → (main, False)."""
    if not value:
        return None, False
    s = str(value).strip().lower()
    if "liên hệ" in s or "corresponding" in s:
        return "corresponding", True
    if s in ("đtg", "dtg") or "đồng tác giả" in s or "đồng" in s:
        return "co", False
    if s in ("tg",) or "tác giả" in s:
        return "main", False
    return None, False


def split_members(value) -> list[str]:
    """"A, B; C\\nD" → ["A","B","C","D"] (tách phẩy/chấm phẩy/xuống dòng, bỏ rỗng)."""
    if not value:
        return []
    parts = re.split(r"[,;\n]", str(value))
    return [p.strip() for p in parts if p.strip()]


# Tiền tố học hàm/học vị hay đứng trước tên user ("ThS. Nguyễn A") — bỏ khi so khớp với
# tên trơn trong Excel ("Nguyễn A").
_TITLE_PREFIX = re.compile(
    r"^(gs\.?ts|pgs\.?ts|gs|pgs|ts|ths|thac si|tien si|bs|ks|cn|ncs|cvc|cv)\.?\s+",
    re.I,
)


def normalize_name(value) -> str:
    """Chuẩn hoá tên để so khớp: bỏ học hàm/học vị, bỏ dấu, thường hoá, gộp khoảng trắng."""
    if not value:
        return ""
    s = unicodedata.normalize("NFD", str(value))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = re.sub(r"\s+", " ", s).strip().lower()
    # bỏ tiền tố học hàm/học vị (lặp phòng khi có nhiều: "gs ts ...")
    prev = None
    while prev != s:
        prev = s
        s = _TITLE_PREFIX.sub("", s)
    return s.strip()


# ---------------------------------------------------------------- DB import (side effects)

def _build_name_index(db):
    from app.models.user import User

    idx = {}
    for u in db.query(User).all():
        idx[normalize_name(u.full_name)] = u.id
    return idx


def resolve_person(name: str, name_index: dict):
    """Trả (user_id, external_name): khớp tên → user_id; không khớp → external_name."""
    key = normalize_name(name)
    uid = name_index.get(key)
    if uid:
        return uid, None
    return None, name.strip()


def run_import(path: str, academic_year: str, dry_run: bool, only=None) -> dict:
    import openpyxl

    only = set(only) if only else None  # None = import tất cả; hoặc tập tên section
    def want(section: str) -> bool:
        return only is None or section in only

    wb = openpyxl.load_workbook(path, data_only=True)
    report = {"projects": 0, "publications": 0, "contracts": 0, "teaching": 0,
              "staff_activities": 0, "certificates": 0, "external_people": 0}

    db = None
    name_index = {}
    if not dry_run:
        from app.db.database import SessionLocal

        db = SessionLocal()
        name_index = _build_name_index(db)

    def cell(ws, r, c):
        v = ws.cell(r, c).value
        return v if v not in (None, "") else None

    # ---- NCKH: nhiệm vụ/đề tài (header r7, data từ r8) ----
    ws = wb["NCKH"]
    from app.models.research import (
        Publication,
        PublicationAuthor,
        ProjectMember,
        ResearchContract,
        ResearchProject,
    )

    for r in (range(8, 18) if want("projects") else range(0)):
        title = cell(ws, r, 2)
        if not title:
            continue
        y0, y1 = parse_year_range(cell(ws, r, 6))
        report["projects"] += 1
        if not dry_run:
            lead_raw = cell(ws, r, 4)
            lead_uid, lead_ext = (None, None)
            if lead_raw:
                lead_uid, lead_ext = resolve_person(str(lead_raw), name_index)
            if lead_uid is None and not lead_ext:
                lead_ext = "(Chưa cập nhật)"  # nguồn thiếu chủ nhiệm — giữ ràng buộc lead-present
            proj = ResearchProject(
                title=str(title).strip(),
                lead_user_id=lead_uid, lead_external_name=lead_ext,
                academic_year=academic_year,
                budget_amount=parse_budget(cell(ws, r, 7)),
                start_date=None, end_date=None,
            )
            db.add(proj)
            db.flush()
            for nm in split_members(cell(ws, r, 5)):
                uid, ext = resolve_person(nm, name_index)
                if ext:
                    report["external_people"] += 1
                db.add(ProjectMember(project_id=proj.id, user_id=uid, external_name=ext))

    # ---- NCKH: bài báo trong nước (r22+), quốc tế (r41+), hội nghị (r55+) ----
    def import_pub(rows, *, scope, is_conf, indexing_cols=None, author_role_col=4, members_col=5,
                   journal_col=3, year_col=6, title_col=2):
        for r in rows:
            title = cell(ws, r, title_col)
            if not title:
                continue
            report["publications"] += 1
            if dry_run:
                continue
            idx = {"is_scie": False, "is_ssci": False, "is_scopus": False, "is_aci": False}
            if indexing_cols:
                idx = parse_indexing(cell(ws, r, indexing_cols[0]), cell(ws, r, indexing_cols[1]), cell(ws, r, indexing_cols[2]))
            pub = Publication(
                title=str(title).strip(), journal=str(cell(ws, r, journal_col) or "") or None,
                year=parse_int(cell(ws, r, year_col)), academic_year=academic_year,
                type="conference" if is_conf else "paper", pub_scope=scope, **idx,
            )
            db.add(pub)
            db.flush()
            role, is_corr = parse_author_role(cell(ws, r, author_role_col))
            order = 1
            for nm in split_members(cell(ws, r, members_col)):
                uid, ext = resolve_person(nm, name_index)
                if ext:
                    report["external_people"] += 1
                db.add(PublicationAuthor(
                    publication_id=pub.id, author_order=order, user_id=uid,
                    external_name=ext if uid is None else None,
                    is_corresponding=(order == 1 and is_corr), author_role=role if order == 1 else "co",
                ))
                order += 1

    if want("publications"):
        import_pub(range(22, 37), scope="domestic", is_conf=False, author_role_col=4, members_col=5)
        import_pub(range(41, 52), scope="international", is_conf=False, indexing_cols=(9, 10, 11),
                   author_role_col=6, members_col=7, year_col=8)
        import_pub(range(55, 72), scope=None, is_conf=True, author_role_col=6, members_col=6,
                   journal_col=3, year_col=7)

    # ---- NCKH: hợp đồng (r75+) ----
    for r in (range(75, 83) if want("contracts") else range(0)):
        title = cell(ws, r, 2)
        if not title:
            continue
        report["contracts"] += 1
        if not dry_run:
            db.add(ResearchContract(
                title=str(title).strip(), contract_type=str(cell(ws, r, 3) or "") or None,
                value_amount=parse_budget(cell(ws, r, 6)), partner_org=str(cell(ws, r, 7) or "") or None,
                academic_year=academic_year,
            ))

    # ---- ĐÀO TẠO: môn giảng dạy (data r8+, header 2 tầng r6-7) ----
    from app.models.research import TeachingCourse

    wt = wb["ĐÀO TẠO"]
    current_user_name = None
    _TEACH_SKIP = ("họ tên", "công tác đào tạo")  # dòng tiêu đề/nhóm, không phải giảng viên
    for r in (range(8, wt.max_row + 1) if want("teaching") else range(0)):
        name = wt.cell(r, 2).value
        course = wt.cell(r, 3).value
        if name and str(name).strip().lower() not in _TEACH_SKIP and "công tác đào tạo" not in str(name).lower():
            current_user_name = str(name).strip()
        if not course:
            continue
        report["teaching"] += 1
        if not dry_run and current_user_name:
            uid, ext = resolve_person(current_user_name, name_index)
            if ext:
                report["external_people"] += 1
            db.add(TeachingCourse(
                user_id=uid, lecturer_external_name=ext if uid is None else None,
                course_name=str(course).strip(), academic_year=academic_year,
                hk1_theory_hours=parse_int(wt.cell(r, 4).value), hk1_practice_hours=parse_int(wt.cell(r, 5).value),
                hk2_theory_hours=parse_int(wt.cell(r, 6).value), hk2_practice_hours=parse_int(wt.cell(r, 7).value),
                note=str(wt.cell(r, 8).value or "") or None,
            ))

    # ---- CÔNG TÁC KHÁC: Đảng/Công đoàn/VILAS ----
    from app.models.research import StaffActivity

    wk = wb["CÔNG TÁC KHÁC"]
    kind_by_title = {"đảng": "dang", "công đoàn": "cong_doan", "vilas": "vilas"}
    cur_kind = "khac"
    for r in (range(1, wk.max_row + 1) if want("staff_activities") else range(0)):
        a = wk.cell(r, 1).value
        b = wk.cell(r, 2).value
        if a and isinstance(a, str) and any(k in a.lower() for k in kind_by_title):
            for k, v in kind_by_title.items():
                if k in a.lower():
                    cur_kind = v
            continue
        if a and str(a).strip().upper() == "STT":
            continue
        if b:  # dòng hoạt động
            report["staff_activities"] += 1
            if not dry_run:
                db.add(StaffActivity(kind=cur_kind, content=str(b).strip(), academic_year=academic_year))

    # ---- PHỤC VỤ CỘNG ĐỒNG: cấp GCN (bảng r14+) ----
    from app.models.research import TrainingCertificate

    wc = wb["PHỤC VỤ CỘNG ĐỒNG"]
    for r in (range(15, wc.max_row + 1) if want("certificates") else range(0)):
        recipient = wc.cell(r, 4).value
        if not recipient:
            continue
        report["certificates"] += 1
        if not dry_run:
            db.add(TrainingCertificate(
                recipient_name=str(recipient).strip(), certificate_no=str(wc.cell(r, 3).value or "") or None,
                course_name=str(wc.cell(r, 5).value or "") or None, academic_year=academic_year,
            ))

    if not dry_run:
        db.commit()
        db.close()
    return report


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--academic-year", default="2024-2025")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--only",
        default=None,
        help="Chỉ import các section (phẩy): projects,publications,contracts,teaching,"
        "staff_activities,certificates. Bỏ trống = tất cả.",
    )
    args = ap.parse_args()
    only = [s.strip() for s in args.only.split(",")] if args.only else None
    rep = run_import(args.file, args.academic_year, args.dry_run, only=only)
    mode = "DRY-RUN (không ghi DB)" if args.dry_run else "ĐÃ GHI DB"
    scope = f" (chỉ: {', '.join(only)})" if only else ""
    print(f"=== Import {mode} — năm học {args.academic_year}{scope} ===")
    for k, v in rep.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    sys.exit(main())
