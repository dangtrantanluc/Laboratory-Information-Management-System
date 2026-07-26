"""RBAC + safety tests cho report_export_service (M6.4).

Không chạm DB thật: mock các service tổng hợp (dashboard/unified_report/system_access)
và audit_service. Mục tiêu: khóa lại hợp đồng RBAC (office/staff/admin/leader) +
formula-injection sanitization — không phải test lại logic của service tổng hợp
(đã có resolve_scope_department/strip_price_fields ở report_common).
"""
import io
import uuid
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

from app.core.deps import CurrentUser
from app.core.exceptions import AppException
from app.services import report_export_service as svc


def _user(role: str, department_id=None) -> CurrentUser:
    return CurrentUser(
        id=uuid.uuid4(),
        email=f"{role}@ribe.vn",
        full_name="Nguyễn Văn A",
        role=role,
        department_id=department_id,
        is_dept_lead=False,
        is_quality_manager=False,
        status="active",
        jti="test-jti",
        token_exp=9999999999,
    )


@pytest.fixture
def db():
    return MagicMock()


# ===================== RBAC gate =====================

def test_export_xlsx_office_samples_403(db):
    with pytest.raises(AppException) as exc:
        svc.export_xlsx(
            db, user=_user("office"), report_type="samples", params={},
            correlation_id=None, ip=None,
        )
    assert exc.value.http_status == 403


def test_export_xlsx_staff_system_access_403(db):
    with patch("app.services.report_common.has_permission", return_value=False):
        with pytest.raises(AppException) as exc:
            svc.export_xlsx(
                db, user=_user("staff"), report_type="system-access", params={},
                correlation_id=None, ip=None,
            )
    assert exc.value.http_status == 403


def test_export_xlsx_admin_system_access_200(db):
    with patch("app.services.report_common.has_permission", return_value=True), \
         patch.object(svc, "system_access_service") as mock_sa, \
         patch.object(svc, "audit_service") as mock_audit:
        mock_sa.system_access.return_value = ({"totals": {}, "timeline": []}, {})
        content, filename = svc.export_xlsx(
            db, user=_user("admin"), report_type="system-access", params={},
            correlation_id="cid-1", ip="127.0.0.1",
        )
    assert filename == "bao-cao-system-access.xlsx"
    assert content  # bytes non-empty
    mock_audit.log_action.assert_called_once()
    assert mock_audit.log_action.call_args.kwargs["action"] == "REPORT_EXPORT"
    db.commit.assert_called_once()


def test_export_xlsx_unsupported_report_type_404(db):
    with pytest.raises(AppException) as exc:
        svc.export_xlsx(
            db, user=_user("admin"), report_type="bogus", params={},
            correlation_id=None, ip=None,
        )
    assert exc.value.http_status == 404
    assert exc.value.code == "REPORT_TYPE_NOT_FOUND"


def test_export_pdf_unsupported_type_422(db):
    with pytest.raises(AppException) as exc:
        svc.export_pdf(
            db, user=_user("admin"), report_type="samples", params={},
            correlation_id=None, ip=None,
        )
    assert exc.value.http_status == 422
    assert exc.value.code == "PDF_NOT_SUPPORTED"


# ===================== Formula-injection sanitization =====================

@pytest.mark.parametrize("raw,expected", [
    ("=cmd|'/c calc'!A1", "'=cmd|'/c calc'!A1"),
    ("+1+1", "'+1+1"),
    ("-1", "'-1"),
    ("@SUM(A1)", "'@SUM(A1)"),
    ("bình thường", "bình thường"),
    (42, 42),
    (None, None),
])
def test_safe_cell_sanitizes_formula_triggers(raw, expected):
    assert svc._safe_cell(raw) == expected


def test_export_xlsx_dashboard_sanitizes_malicious_cell_values(db):
    malicious = {
        "scope": {"department_name": "Phòng A"},
        "samples": {"=HYPERLINK(\"http://evil\")": 1, "total": 10},
    }
    with patch.object(svc, "dashboard_service") as mock_dash, \
         patch.object(svc, "audit_service"):
        mock_dash.get_dashboard.return_value = (malicious, {})
        mock_dash.get_charts.return_value = ({}, {})
        content, _ = svc.export_xlsx(
            db, user=_user("admin"), report_type="dashboard", params={},
            correlation_id=None, ip=None,
        )

    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    values = [cell for row in ws.iter_rows() for cell in row if cell.value]
    texts = [str(c.value) for c in values]
    assert any(t.startswith("'=HYPERLINK") for t in texts)
    assert not any(t.startswith("=") for t in texts)  # không còn cell nào là formula sống


# ===================== Chemicals cost column (upstream-driven) =====================

def test_write_sheet_chemicals_includes_cost_column_only_when_present():
    from openpyxl import Workbook

    ws = Workbook().active
    data_with_cost = {"by_measurement_group": [
        {"measurement_group": "mass", "base_unit": "g", "total_qty": 100, "consumption_cost": 5000},
    ]}
    svc._write_sheet(ws, "chemicals", data_with_cost)
    header_row = [c.value for c in ws[2]]
    assert "Chi phí" in header_row


def test_write_sheet_chemicals_excludes_cost_column_when_absent():
    from openpyxl import Workbook

    ws = Workbook().active
    data_no_cost = {"by_measurement_group": [
        {"measurement_group": "mass", "base_unit": "g", "total_qty": 100},
    ]}
    svc._write_sheet(ws, "chemicals", data_no_cost)
    header_row = [c.value for c in ws[2]]
    assert "Chi phí" not in header_row
