"""Xuất BẢNG BÁO GIÁ ra Excel — dựng đúng layout file mẫu của Viện (m29/GĐ4).

Layout mẫu (1792026 0722 Báo gía - CÔNG TY TNHH BEJO VIỆT NAM.xls):
  header Trường/Viện + Mã số/Web/Mail/Tel · "BẢNG BÁO GIÁ" · Kính gửi/Địa chỉ/Email/Tel
  · câu mở đầu · bảng 6 cột (STT · Loại/Tên mẫu · Chỉ tiêu · Số lượng · Đơn giá · Thành tiền)
  · Cộng / VAT x% / Tổng cộng · Ghi chú · ngày + Người lập.
Tên mẫu chỉ in ở DÒNG ĐẦU mỗi nhóm (giống mẫu).
"""
import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="000000")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
MONEY = "#,##0"

ORG_LINE1 = "TRƯỜNG ĐẠI HỌC NÔNG LÂM TP. HỒ CHÍ MINH"
ORG_LINE2 = "VIỆN NGHIÊN CỨU CÔNG NGHỆ SINH HỌC VÀ MÔI TRƯỜNG"
INTRO = (
    "     Theo yêu cầu của Quý Khách hàng, Viện Nghiên cứu Công nghệ Sinh học và Môi trường "
    "xin gửi đến Quý khách hàng bảng báo giá phân tích như sau:"
)
NOTES = (
    "     Bảng báo giá này có giá trị trong vòng 1 tháng.",
    "     Thời gian trả kết quả: tùy thuộc vào chỉ tiêu và số lượng mẫu khách hàng gửi, "
    "trường hợp đặc biệt sẽ được thương lượng cụ thể.",
    "     Địa chỉ gửi mẫu (trực tiếp hoặc bưu điện): Phòng 211 - Phòng nhận mẫu, Tòa nhà A2",
)
HEADERS = ("STT", "Loại/ Tên mẫu", "Chỉ tiêu thử nghiệm", "Số lượng", "Đơn giá\n(VNĐ)", "Thành tiền\n(VNĐ)")


def build_xlsx(quotation: dict) -> bytes:
    """quotation: dict từ quotation_service._serialize (kèm items)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo giá"

    widths = (6, 22, 46, 10, 16, 18)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    # --- Header đơn vị (trái) + liên hệ (phải) ---
    ws.cell(r, 2, ORG_LINE1).font = Font(bold=True, size=11)
    ws.cell(r, 5, f"Mã số: {quotation.get('code') or ''}").alignment = LEFT
    r += 1
    ws.cell(r, 2, ORG_LINE2).font = Font(bold=True, size=11, color="1A6E4A")
    ws.cell(r, 5, "Web: ribe.hcmuaf.edu.vn")
    r += 1
    ws.cell(r, 5, "Mail: ptm.ribe@hcmuaf.edu.vn")
    r += 1
    ws.cell(r, 5, "Tel: 028 3724 6019")
    r += 2

    # --- Tiêu đề ---
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, "BẢNG BÁO GIÁ")
    c.font = Font(bold=True, size=16)
    c.alignment = CENTER
    r += 2

    # --- Kính gửi ---
    for label, key in (
        ("Kính gửi:", "customer_name"),
        ("Địa chỉ:", "customer_address"),
        ("Email:", "customer_email"),
        ("Tel:", "customer_phone"),
    ):
        ws.cell(r, 1, label).font = Font(bold=(key == "customer_name"))
        v = quotation.get(key) or ""
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        cell = ws.cell(r, 2, v)
        cell.alignment = LEFT
        if key == "customer_name":
            cell.font = Font(bold=True)
        r += 1
    r += 1

    # --- Câu mở đầu ---
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(r, 1, INTRO).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[r].height = 30
    r += 2

    # --- Bảng chi tiết ---
    head_row = r
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(head_row, i, h)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.border = BOX
        cell.fill = PatternFill("solid", fgColor="EAF3EE")
    ws.row_dimensions[head_row].height = 30
    r += 1

    items = quotation.get("items") or []
    prev_sample = None
    for i, it in enumerate(items, start=1):
        sample = it.get("sample_name") or ""
        # Chỉ in tên mẫu ở dòng ĐẦU mỗi nhóm (đúng như file mẫu)
        show = sample if sample != prev_sample else ""
        prev_sample = sample
        qty = int(it.get("quantity") or 1)
        price = Decimal(str(it.get("unit_price") or 0))
        amount = Decimal(str(it.get("amount") or (price * qty)))

        vals = (i, show, it.get("parameter_name") or "", qty, float(price), float(amount))
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(r, ci, v)
            cell.border = BOX
            if ci in (1, 4):
                cell.alignment = CENTER
            elif ci in (5, 6):
                cell.alignment = RIGHT
                cell.number_format = MONEY
            else:
                cell.alignment = LEFT
        r += 1

    # --- Cộng / VAT / Tổng cộng ---
    def total_row(label: str, value, bold: bool = False):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        lc = ws.cell(r, 1, label)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.font = Font(bold=bold)
        lc.border = BOX
        for ci in range(2, 6):
            ws.cell(r, ci).border = BOX
        vc = ws.cell(r, 6, float(Decimal(str(value or 0))))
        vc.number_format = MONEY
        vc.alignment = RIGHT
        vc.font = Font(bold=bold)
        vc.border = BOX
        r += 1

    # Định dạng % gọn: 8.00 → "8", 10.50 → "10.5" (KHÔNG dùng normalize() vì ra 1E+1)
    vat_rate = Decimal(str(quotation.get("vat_rate") or 0))
    vat_txt = f"{vat_rate:f}".rstrip("0").rstrip(".") or "0"
    total_row("Cộng:", quotation.get("subtotal"))
    total_row(f"VAT {vat_txt}%:", quotation.get("vat_amount"))
    total_row("Tổng cộng:", quotation.get("total"), bold=True)
    r += 1

    # --- Ghi chú ---
    ws.cell(r, 1, "Ghi chú:").font = Font(bold=True)
    r += 1
    for line in NOTES:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1, line).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 26
        r += 1
    if quotation.get("note"):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(r, 1, f"     {quotation['note']}").alignment = LEFT
        r += 1
    r += 2

    # --- Ngày + người lập ---
    issue = quotation.get("issue_date")
    day = month = year = ""
    if issue:
        s = str(issue)
        parts = s.split("-")
        if len(parts) == 3:
            year, month, day = parts[0], parts[1], parts[2][:2]
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(r, 4, f"Tp.HCM, ngày {day} tháng {month} năm {year}").alignment = CENTER
    r += 1
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(r, 4, "Người lập").font = Font(bold=True)
    ws.cell(r, 4).alignment = CENTER
    r += 3
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(r, 4, quotation.get("created_by_name") or "").alignment = CENTER

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
