"""Service kho biểu mẫu VILAS (GĐ3) — QLCL quản trị template; phòng lab nộp minh chứng.

Luồng: KTV/trưởng phòng tải template về (attachments owner_type='form_template'),
điền, tạo submission + upload file (owner_type='form_submission') → thông báo Phòng QLCL.
"""
import io
import uuid
from typing import Optional

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.error_codes import ErrorCode
from app.core.concurrency import export_slot
from app.core.deps import CurrentUser
from app.core.exceptions import AppException, not_found
from app.models.attachment import Attachment
from app.models.department import Department
from app.models.form import FormSubmission, FormTemplate
from app.models.user import User
from app.services import audit_read_service, audit_service, notification_service

EXPORT_MAX_ROWS = 10000


def _forbidden(msg: str = "Bạn không có quyền thực hiện thao tác này") -> AppException:
    return AppException(ErrorCode.FORBIDDEN, msg, 403)


def _files_of(db: Session, owner_type: str, owner_id: uuid.UUID) -> list[dict]:
    rows = db.execute(
        select(Attachment)
        .where(
            Attachment.owner_type == owner_type,
            Attachment.owner_id == owner_id,
            Attachment.deleted_at.is_(None),
        )
        .order_by(Attachment.uploaded_at.desc())
    ).scalars().all()
    return [
        {"id": a.id, "file_name": a.file_name, "mime": a.mime, "size": a.size,
         "uploaded_at": a.uploaded_at}
        for a in rows
    ]


def _serialize_template(db: Session, t: FormTemplate) -> dict:
    return {
        "id": t.id,
        "code": t.code,
        "title": t.title,
        "iso_clause": t.iso_clause,
        "category": t.category,
        "year": t.year,
        "is_active": t.is_active,
        "note": t.note,
        "files": _files_of(db, "form_template", t.id),
        "created_at": t.created_at,
    }


# ===== Templates =====
def list_templates(
    db: Session,
    *,
    iso_clause: Optional[str],
    category: Optional[str],
    q: Optional[str],
    active_only: bool,
    page: int,
    limit: int,
) -> tuple[list[dict], int]:
    conds = []
    if active_only:
        conds.append(FormTemplate.is_active.is_(True))
    if iso_clause:
        conds.append(FormTemplate.iso_clause == iso_clause)
    if category:
        conds.append(FormTemplate.category == category)
    if q:
        like = f"%{q.strip()}%"
        conds.append(FormTemplate.title.ilike(like) | FormTemplate.code.ilike(like))
    total = db.execute(
        select(func.count()).select_from(FormTemplate).where(*conds)
    ).scalar_one()
    rows = db.execute(
        select(FormTemplate)
        .where(*conds)
        .order_by(FormTemplate.iso_clause, FormTemplate.code)
        .offset((page - 1) * limit)
        .limit(limit)
    ).scalars().all()
    return [_serialize_template(db, t) for t in rows], total


def create_template(
    db: Session,
    *,
    user: CurrentUser,
    code: str,
    title: str,
    iso_clause: str,
    category: str,
    year: Optional[int],
    note: Optional[str],
    correlation_id: Optional[str],
    ip: Optional[str],
) -> dict:
    dup = db.execute(
        select(FormTemplate.id).where(FormTemplate.code == code.strip())
    ).scalar_one_or_none()
    if dup is not None:
        raise AppException(ErrorCode.DUPLICATE_CODE, "Mã biểu mẫu đã tồn tại", 409)
    t = FormTemplate(
        code=code.strip(),
        title=title.strip(),
        iso_clause=iso_clause.strip(),
        category=category,
        year=year,
        note=note,
        created_by=user.id,
        updated_by=user.id,
    )
    db.add(t)
    db.flush()
    audit_service.log_action(
        db, action="FORM_TEMPLATE_CREATE", resource="form_template",
        user_id=user.id, resource_id=t.id, correlation_id=correlation_id, ip=ip,
        detail={"code": t.code},
    )
    db.commit()
    db.refresh(t)
    return _serialize_template(db, t)


def update_template(
    db: Session,
    *,
    user: CurrentUser,
    template_id: uuid.UUID,
    changes: dict,
    correlation_id: Optional[str],
    ip: Optional[str],
) -> dict:
    t = db.get(FormTemplate, template_id)
    if t is None:
        raise not_found("Không tìm thấy biểu mẫu")
    for k, v in changes.items():
        setattr(t, k, v)
    t.updated_by = user.id
    audit_service.log_action(
        db, action="FORM_TEMPLATE_UPDATE", resource="form_template",
        user_id=user.id, resource_id=t.id, correlation_id=correlation_id, ip=ip,
        detail=changes,
    )
    db.commit()
    db.refresh(t)
    return _serialize_template(db, t)


# ===== Submissions (minh chứng) =====
def _batch_submission_refs(
    db: Session, subs: list[FormSubmission]
) -> tuple[dict, dict, dict, dict]:
    """Batch-load template/department/user/files cho danh sách FormSubmission — tránh N+1
    (PRODUCTION_READINESS_REVIEW §Performance: "form_service.py (_serialize_submission,
    4 db.get() per row)")."""
    tpl_ids = {s.template_id for s in subs}
    tpls_by_id = {
        t.id: t
        for t in (
            db.execute(select(FormTemplate).where(FormTemplate.id.in_(tpl_ids))).scalars()
            if tpl_ids
            else []
        )
    }

    dept_ids = {s.department_id for s in subs if s.department_id}
    depts_by_id = {
        d.id: d
        for d in (
            db.execute(select(Department).where(Department.id.in_(dept_ids))).scalars()
            if dept_ids
            else []
        )
    }

    user_ids = {s.submitted_by for s in subs} | {s.reviewed_by for s in subs if s.reviewed_by}
    users_by_id = {
        u.id: u
        for u in (
            db.execute(select(User).where(User.id.in_(user_ids))).scalars() if user_ids else []
        )
    }

    sub_ids = [s.id for s in subs]
    files_by_owner: dict = {sid: [] for sid in sub_ids}
    if sub_ids:
        att_rows = db.execute(
            select(Attachment)
            .where(
                Attachment.owner_type == "form_submission",
                Attachment.owner_id.in_(sub_ids),
                Attachment.deleted_at.is_(None),
            )
            .order_by(Attachment.uploaded_at.desc())
        ).scalars().all()
        for a in att_rows:
            files_by_owner.setdefault(a.owner_id, []).append(
                {
                    "id": a.id,
                    "file_name": a.file_name,
                    "mime": a.mime,
                    "size": a.size,
                    "uploaded_at": a.uploaded_at,
                }
            )

    return tpls_by_id, depts_by_id, users_by_id, files_by_owner


def _serialize_submission(
    db: Session,
    s: FormSubmission,
    *,
    tpls_by_id: Optional[dict] = None,
    depts_by_id: Optional[dict] = None,
    users_by_id: Optional[dict] = None,
    files_by_owner: Optional[dict] = None,
) -> dict:
    # Không truyền map batch (vd: gọi cho 1 submission đơn lẻ sau create/review) → fallback
    # về db.get() từng cột như cũ, đúng hành vi, chỉ không batch.
    if tpls_by_id is None:
        tpl = db.get(FormTemplate, s.template_id)
        dept = db.get(Department, s.department_id) if s.department_id else None
        submitter = db.get(User, s.submitted_by)
        reviewer = db.get(User, s.reviewed_by) if s.reviewed_by else None
        files = _files_of(db, "form_submission", s.id)
    else:
        tpl = tpls_by_id.get(s.template_id)
        dept = depts_by_id.get(s.department_id) if s.department_id else None
        submitter = users_by_id.get(s.submitted_by)
        reviewer = users_by_id.get(s.reviewed_by) if s.reviewed_by else None
        files = (files_by_owner or {}).get(s.id, [])
    return {
        "id": s.id,
        "template_id": s.template_id,
        "template_code": tpl.code if tpl else None,
        "template_title": tpl.title if tpl else None,
        "department_id": s.department_id,
        "department_name": dept.name if dept else None,
        "year": s.year,
        "note": s.note,
        "submitted_by": s.submitted_by,
        "submitted_by_name": submitter.full_name if submitter else None,
        "submitted_at": s.submitted_at,
        "files": files,
        "status": s.status,
        "reviewed_by": s.reviewed_by,
        "reviewed_by_name": reviewer.full_name if reviewer else None,
        "reviewed_at": s.reviewed_at,
        "reject_reason": s.reject_reason,
    }


def _is_privileged(user: CurrentUser) -> bool:
    return user.role in ("admin", "leader", "qms")


class _NoAccess(Exception):
    """Non-privileged user chưa gán phòng ban — không có gì để thấy."""


def _submission_conditions(
    *,
    user: CurrentUser,
    template_id: Optional[uuid.UUID],
    department_id: Optional[uuid.UUID],
    year: Optional[int],
    status: Optional[str] = None,
) -> list:
    conds = []
    if template_id:
        conds.append(FormSubmission.template_id == template_id)
    if year:
        conds.append(FormSubmission.year == year)
    if status:
        conds.append(FormSubmission.status == status)
    # Phòng lab chỉ thấy minh chứng của phòng mình; QLCL/admin/leader thấy tất cả.
    if not _is_privileged(user):
        if user.department_id is None:
            raise _NoAccess()
        conds.append(FormSubmission.department_id == user.department_id)
    elif department_id:
        conds.append(FormSubmission.department_id == department_id)
    return conds


def list_submissions(
    db: Session,
    *,
    user: CurrentUser,
    template_id: Optional[uuid.UUID],
    department_id: Optional[uuid.UUID],
    year: Optional[int] = None,
    status: Optional[str] = None,
    page: int,
    limit: int,
) -> tuple[list[dict], int]:
    try:
        conds = _submission_conditions(
            user=user, template_id=template_id, department_id=department_id,
            year=year, status=status,
        )
    except _NoAccess:
        return [], 0
    total = db.execute(
        select(func.count()).select_from(FormSubmission).where(*conds)
    ).scalar_one()
    rows = db.execute(
        select(FormSubmission)
        .where(*conds)
        .order_by(FormSubmission.submitted_at.desc())
        .offset((page - 1) * limit)
        .limit(limit)
    ).scalars().all()
    tpls_by_id, depts_by_id, users_by_id, files_by_owner = _batch_submission_refs(db, rows)
    items = [
        _serialize_submission(
            db,
            s,
            tpls_by_id=tpls_by_id,
            depts_by_id=depts_by_id,
            users_by_id=users_by_id,
            files_by_owner=files_by_owner,
        )
        for s in rows
    ]
    return items, total


def export_submissions_xlsx(
    db: Session,
    *,
    user: CurrentUser,
    template_id: Optional[uuid.UUID],
    department_id: Optional[uuid.UUID],
    year: Optional[int],
    correlation_id: Optional[str],
    ip: Optional[str],
) -> bytes:
    with export_slot():
        try:
            conds = _submission_conditions(
                user=user, template_id=template_id, department_id=department_id, year=year
            )
        except _NoAccess:
            conds = None

        total = 0
        rows: list[FormSubmission] = []
        if conds is not None:
            total = db.execute(
                select(func.count()).select_from(FormSubmission).where(*conds)
            ).scalar_one()
            if total > EXPORT_MAX_ROWS:
                raise AppException(
                    ErrorCode.EXPORT_TOO_LARGE,
                    f"Kết quả {total} dòng vượt ngưỡng {EXPORT_MAX_ROWS}. Thu hẹp bộ lọc.",
                    422,
                )
            rows = db.execute(
                select(FormSubmission).where(*conds).order_by(FormSubmission.submitted_at.desc())
            ).scalars().all()

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        COLUMNS = [
            ("STT", 6),
            ("Mã BM", 16),
            ("Tên biểu mẫu", 46),
            ("Phòng", 24),
            ("Năm", 8),
            ("Người nộp", 22),
            ("Thời gian nộp", 18),
            ("Tệp đính kèm", 40),
        ]
        NAVY = "1F3864"
        BAND = "F2F5FA"
        thin = Side(style="thin", color="BFBFBF")
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = "Minh chứng VILAS"

        # ── Tiêu đề + tóm tắt bộ lọc ──
        n_cols = len(COLUMNS)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        title_cell = ws.cell(row=1, column=1, value="TỔNG HỢP MINH CHỨNG BIỂU MẪU VILAS")
        title_cell.font = Font(name="Calibri", size=14, bold=True, color=NAVY)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        filter_bits = [f"Tổng {total} minh chứng"]
        if year:
            filter_bits.append(f"Năm {year}")
        if department_id:
            dept = db.get(Department, department_id)
            if dept:
                filter_bits.append(f"Phòng {dept.name}")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        meta_cell = ws.cell(row=2, column=1, value=" · ".join(filter_bits))
        meta_cell.font = Font(name="Calibri", size=10, italic=True, color="595959")
        meta_cell.alignment = Alignment(horizontal="center")

        # ── Header ──
        header_row = 4
        for col_idx, (label, width) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[header_row].height = 20

        # ── Dữ liệu ──
        center_cols = {1, 5, 7}  # STT, Năm, Thời gian nộp
        tpls_by_id, depts_by_id, users_by_id, files_by_owner = _batch_submission_refs(db, rows)
        for i, s in enumerate(rows, start=1):
            data = _serialize_submission(
                db,
                s,
                tpls_by_id=tpls_by_id,
                depts_by_id=depts_by_id,
                users_by_id=users_by_id,
                files_by_owner=files_by_owner,
            )
            r = header_row + i
            values = [
                i,
                data["template_code"] or "",
                data["template_title"] or "",
                data["department_name"] or "",
                data["year"] or "",
                data["submitted_by_name"] or "",
                data["submitted_at"].strftime("%Y-%m-%d %H:%M") if data["submitted_at"] else "",
                "; ".join(f["file_name"] for f in data["files"]) or "—",
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col_idx, value=value)
                cell.border = cell_border
                cell.alignment = Alignment(
                    horizontal="center" if col_idx in center_cols else "left",
                    vertical="center",
                    wrap_text=col_idx in (3, 8),
                )
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=BAND)

        if rows:
            last_row = header_row + len(rows)
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{last_row}"

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.sheet_view.showGridLines = False

        audit_service.log_action(
            db,
            action="FORM_SUBMISSION_EXPORT_EXCEL",
            resource="form_submission",
            user_id=user.id,
            correlation_id=correlation_id,
            ip=ip,
            detail={"rows": total, "year": year, "department_id": str(department_id) if department_id else None},
        )
        db.commit()

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def create_submission(
    db: Session,
    *,
    user: CurrentUser,
    template_id: uuid.UUID,
    department_id: Optional[uuid.UUID],
    year: Optional[int],
    note: Optional[str],
    correlation_id: Optional[str],
    ip: Optional[str],
) -> dict:
    tpl = db.get(FormTemplate, template_id)
    if tpl is None:
        raise not_found("Không tìm thấy biểu mẫu")

    # Xác định phòng nộp: KTV/trưởng phòng ép phòng mình; admin/QLCL có thể chỉ định.
    if _is_privileged(user):
        dept_id = department_id or user.department_id
    else:
        dept_id = user.department_id
        if department_id is not None and department_id != user.department_id:
            raise _forbidden("Bạn chỉ được nộp minh chứng cho phòng của mình")
    if dept_id is None:
        raise AppException(ErrorCode.VALIDATION_ERROR, "Cần chỉ định phòng nộp minh chứng", 400)

    sub = FormSubmission(
        template_id=template_id,
        department_id=dept_id,
        year=year,
        note=note,
        submitted_by=user.id,
    )
    db.add(sub)
    db.flush()

    # Thông báo cho toàn bộ nhân sự Phòng QLCL (role qms) — "QLCL nhận thông báo khi up".
    dept = db.get(Department, dept_id)
    dept_name = dept.name if dept else ""
    qms_users = db.execute(
        select(User.id).where(User.role == "qms", User.status == "active")
    ).scalars().all()
    for qid in qms_users:
        notification_service.create_notification(
            db,
            user_id=qid,
            type="FORM_SUBMITTED",
            title="Có minh chứng biểu mẫu mới",
            body=f"{tpl.code} — {tpl.title} · {dept_name}"
            + (f" ({year})" if year else ""),
            ref_type="form_submission",
            ref_id=sub.id,
        )

    audit_service.log_action(
        db, action="FORM_SUBMISSION_CREATE", resource="form_submission",
        user_id=user.id, resource_id=sub.id, correlation_id=correlation_id, ip=ip,
        detail={"template": tpl.code, "department_id": str(dept_id)},
    )
    db.commit()
    db.refresh(sub)
    return _serialize_submission(db, sub)


# ===== Duyệt minh chứng (M21) =====
def _get_submission_or_404(db: Session, submission_id: uuid.UUID) -> FormSubmission:
    sub = db.get(FormSubmission, submission_id)
    if sub is None:
        raise not_found("Không tìm thấy minh chứng")
    return sub


def approve_submission(
    db: Session,
    *,
    user: CurrentUser,
    submission_id: uuid.UUID,
    correlation_id: Optional[str],
    ip: Optional[str],
) -> dict:
    sub = _get_submission_or_404(db, submission_id)
    if sub.status != "pending":
        raise AppException(
            ErrorCode.INVALID_STATE,
            f"Minh chứng đã được xử lý (trạng thái hiện tại: '{sub.status}')",
            422,
        )
    sub.status = "approved"
    sub.reviewed_by = user.id
    sub.reviewed_at = func.now()

    audit_service.log_action(
        db, action="FORM_SUBMISSION_APPROVE", resource="form_submission",
        user_id=user.id, resource_id=sub.id, correlation_id=correlation_id, ip=ip,
        detail={},
    )
    tpl = db.get(FormTemplate, sub.template_id)
    notification_service.create_notification(
        db,
        user_id=sub.submitted_by,
        type="FORM_SUBMISSION_APPROVED",
        title="Minh chứng đã được duyệt",
        body=f"{tpl.code if tpl else ''} — {tpl.title if tpl else ''}",
        ref_type="form_submission",
        ref_id=sub.id,
    )
    db.commit()
    db.refresh(sub)
    return _serialize_submission(db, sub)


def reject_submission(
    db: Session,
    *,
    user: CurrentUser,
    submission_id: uuid.UUID,
    reject_reason: str,
    correlation_id: Optional[str],
    ip: Optional[str],
) -> dict:
    sub = _get_submission_or_404(db, submission_id)
    if sub.status != "pending":
        raise AppException(
            ErrorCode.INVALID_STATE,
            f"Minh chứng đã được xử lý (trạng thái hiện tại: '{sub.status}')",
            422,
        )
    sub.status = "rejected"
    sub.reviewed_by = user.id
    sub.reviewed_at = func.now()
    sub.reject_reason = reject_reason

    audit_service.log_action(
        db, action="FORM_SUBMISSION_REJECT", resource="form_submission",
        user_id=user.id, resource_id=sub.id, correlation_id=correlation_id, ip=ip,
        detail={"reject_reason": reject_reason},
    )
    tpl = db.get(FormTemplate, sub.template_id)
    notification_service.create_notification(
        db,
        user_id=sub.submitted_by,
        type="FORM_SUBMISSION_REJECTED",
        title="Minh chứng bị từ chối",
        body=f"{tpl.code if tpl else ''} — {tpl.title if tpl else ''}: {reject_reason}",
        ref_type="form_submission",
        ref_id=sub.id,
    )
    db.commit()
    db.refresh(sub)
    return _serialize_submission(db, sub)


def submission_history(db: Session, *, submission_id: uuid.UUID) -> list[dict]:
    """Lịch sử duyệt — tái dùng bảng audit_logs chung, lọc theo resource_id."""
    _get_submission_or_404(db, submission_id)
    items, _total = audit_read_service.list_audit_logs(
        db,
        user_id=None, action=None, resource="form_submission", resource_id=submission_id,
        correlation_id=None, date_from=None, date_to=None, ip=None,
        page=1, limit=50,
    )
    return items
